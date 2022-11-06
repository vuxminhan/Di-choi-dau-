from openpyxl import  load_workbook
import pandas as pd
wb = load_workbook(r'data1/DataFrame.xlsx')
ws = wb.active
# def create_eachgame_in4():
def take_number_rating(m):
    a = []
    for i in m:
        ma = i.split()
        for l in ma[0]:
            if l == '.':
                ma[0] = ma[0].replace(l,'')
        a.append(int(ma[0]))
    return a
def take_name_game():
    e = []
    for col in range(1,2700):
        if ws['F' + str(col)].value != ws['F' + str(col+ 1)].value:
            e.append(ws['F' + str(col)].value)
        if col == 2699:
            e.append(ws['F' + str(col)].value)
    return e
def swap(list1,a,b):
    list1[a], list1[b] = list1[b], list1[a]
    return list
def append_infor(a,m,n):
    for col in range(1,2763):
        if ws['F' + str(col)].value == m:
            n.append(ws[str(a) + str(col)].value)
    return n
class game_choice():
    def __init__(self,game):
        self.game = game
        self.infor = {}
        self.infor['Name'] = []
        self.infor['Location'] = []
        self.infor['Open_time'] = []
        self.infor['Rating'] = []
        self.infor['Number_of_rating'] = []
        # bien de display the best choice for game
        self.name_place = str()
        self.map_location = []
        self.longtitude = 0
        self.lattitude = 0
        # picture
        self.picture = []
    def add_infor(self):
        self.infor['Name'] = append_infor('A',self.game,self.infor['Name'])
        self.infor['Location'] = append_infor('C',self.game,self.infor['Location'])
        self.infor['Open_time'] = append_infor('B',self.game,self.infor['Open_time'])
        #rating
        rat = []
        rat = append_infor('D', self.game, rat)
        for i in rat:
            for j in i:
                if j == ',':
                    i = i.replace(j,'.')
            self.infor['Rating'].append(i)
        self.infor['Rating'] = [float(i) for i in self.infor['Rating']]
        #number of rating
        self.infor['Number_of_rating'] = append_infor('E', self.game, self.infor['Number_of_rating'])
        self.infor['Number_of_rating'] = take_number_rating(self.infor['Number_of_rating'])
        #location in map for the best choice
        self.map_location = append_infor('G',self.game,self.map_location)
        for i in self.map_location:
            if i != None:
                m = self.map_location.index(i)
                self.name_place = self.infor['Name'][m]
                a = i.split(',')
                self.lattitude = float(a[0])
                self.longtitude = float(a[1])
        #lay link anh
        self.picture = append_infor('H',self.game,self.picture)
    def take_best_choice(self):
        fr = pd.DataFrame.from_dict(self.infor)
        best = fr[fr['Name'] == f'{self.name_place}']
        return best

if __name__ == "__main__":
    a = take_name_game()
    a.pop(0)
    print(a)
    gamei = []
    for i in a:
        gamei.append(game_choice(i))
    for i in range(len(gamei)-1):
        gamei[i].add_infor()
    for i in range(len(gamei)):
        print(gamei[i].name_place)






